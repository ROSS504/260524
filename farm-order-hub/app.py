"""
Farm Order Hub - Web 界面
"""
import json
from datetime import datetime
from pathlib import Path

from flask import (
    Flask, flash, jsonify, redirect, render_template, request,
    send_file, url_for,
)
from openpyxl import load_workbook

from farm_order_hub.storage import Storage

app = Flask(__name__)
app.secret_key = "farm-order-hub-secret"

BASE_DIR = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "output"
DB_PATH = BASE_DIR / "data" / "app.db"

UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

storage = Storage(DB_PATH)


# ══════════════════════════════════════
#  页面
# ══════════════════════════════════════

@app.route("/")
def index():
    types = storage.get_all_types()
    # 每个类型附加统计
    for t in types:
        datasets = storage.get_datasets_by_type(t["type_id"])
        rules = storage.get_rules_by_type(t["type_id"], only_enabled=True)
        t["dataset_count"] = len(datasets)
        t["rule_count"] = len(rules)
    return render_template("index.html", types=types)


@app.route("/type/<int:type_id>")
def type_detail(type_id):
    dtype = storage.get_type(type_id)
    if not dtype:
        flash("类型不存在", "error")
        return redirect(url_for("index"))
    datasets = storage.get_datasets_by_type(type_id)
    rules = storage.get_rules_by_type(type_id)
    messages = storage.get_chat_messages(type_id)
    return render_template("type_detail.html", dtype=dtype, datasets=datasets, rules=rules, messages=messages)


@app.route("/dataset/<int:dataset_id>")
def dataset_detail(dataset_id):
    ds = storage.get_dataset(dataset_id)
    if not ds:
        flash("数据集不存在", "error")
        return redirect(url_for("index"))
    dtype = storage.get_type(ds["type_id"])
    rows = storage.get_data_rows(dataset_id)
    headers = json.loads(ds["raw_headers"])
    # 解析每行数据
    parsed_rows = []
    for r in rows:
        parsed_rows.append({
            "row_id": r["row_id"],
            "data": json.loads(r["row_data"]),
            "processed": r["processed"],
            "result": r["result"],
        })
    return render_template("dataset_detail.html", ds=ds, dtype=dtype, headers=headers, rows=parsed_rows)


# ══════════════════════════════════════
#  操作
# ══════════════════════════════════════

@app.route("/upload", methods=["POST"])
def upload():
    """上传 Excel，选择或新建类型"""
    file = request.files.get("file")
    if not file or not file.filename.endswith((".xlsx", ".xls")):
        flash("请上传 .xlsx 文件", "error")
        return redirect(url_for("index"))

    type_name = request.form.get("type_name", "").strip()
    type_id_str = request.form.get("type_id", "").strip()

    # 确定类型
    if type_name:
        existing = storage.get_type_by_name(type_name)
        if existing:
            type_id = existing["type_id"]
        else:
            type_id = storage.create_type(type_name)
    elif type_id_str:
        type_id = int(type_id_str)
    else:
        flash("请选择或输入数据类型", "error")
        return redirect(url_for("index"))

    # 保存文件
    save_path = UPLOAD_DIR / file.filename
    file.save(str(save_path))

    # 读 Excel
    try:
        wb = load_workbook(str(save_path), read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(all_rows) < 2:
            flash("Excel 至少需要表头+1行数据", "error")
            return redirect(url_for("index"))

        headers = [str(h or f"列{i+1}") for i, h in enumerate(all_rows[0])]
        data_rows = []
        for row in all_rows[1:]:
            if not any(row):
                continue
            row_dict = {}
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                row_dict[h] = str(val) if val is not None else ""
            data_rows.append(row_dict)

        dataset_id = storage.create_dataset(type_id, file.filename, len(data_rows), headers)
        storage.add_data_rows(dataset_id, data_rows)

        flash(f"导入成功：{len(data_rows)} 行数据，已绑定类型「{storage.get_type(type_id)['name']}」", "success")
        return redirect(url_for("type_detail", type_id=type_id))

    except Exception as e:
        flash(f"导入失败: {e}", "error")
        return redirect(url_for("index"))


@app.route("/type/<int:type_id>/chat", methods=["POST"])
def chat_send(type_id):
    """用户在对话窗口发送消息"""
    content = request.form.get("content", "").strip()
    if content:
        storage.add_chat_message(type_id, "user", content)
    return redirect(url_for("type_detail", type_id=type_id) + "#chat")


@app.route("/type/<int:type_id>/save-rule", methods=["POST"])
def save_rule(type_id):
    """从对话中确认一条规则并保存"""
    content = request.form.get("content", "").strip()
    if content:
        storage.add_rule(type_id, content)
        flash("规则已保存", "success")
    return redirect(url_for("type_detail", type_id=type_id) + "#rules")


@app.route("/type/<int:type_id>/rule/<int:rule_id>/toggle", methods=["POST"])
def toggle_rule(type_id, rule_id):
    storage.toggle_rule(rule_id)
    return redirect(url_for("type_detail", type_id=type_id) + "#rules")


@app.route("/type/<int:type_id>/rule/<int:rule_id>/delete", methods=["POST"])
def delete_rule(type_id, rule_id):
    storage.delete_rule(rule_id)
    flash("规则已删除", "success")
    return redirect(url_for("type_detail", type_id=type_id) + "#rules")


@app.route("/type/<int:type_id>/clear-chat", methods=["POST"])
def clear_chat(type_id):
    storage.clear_chat(type_id)
    return redirect(url_for("type_detail", type_id=type_id) + "#chat")


# ══════════════════════════════════════
#  API（供 Claude Code 调用）
# ══════════════════════════════════════

@app.route("/api/types", methods=["GET"])
def api_types():
    return jsonify(storage.get_all_types())


@app.route("/api/type/<int:type_id>/rules", methods=["GET"])
def api_rules(type_id):
    rules = storage.get_rules_by_type(type_id, only_enabled=True)
    return jsonify([{"rule_id": r["rule_id"], "content": r["content"]} for r in rules])


@app.route("/api/type/<int:type_id>/chat", methods=["GET"])
def api_chat_get(type_id):
    return jsonify(storage.get_chat_messages(type_id))


@app.route("/api/type/<int:type_id>/chat", methods=["POST"])
def api_chat_post(type_id):
    """Claude Code 回复消息"""
    data = request.get_json()
    content = data.get("content", "")
    role = data.get("role", "assistant")
    if content:
        storage.add_chat_message(type_id, role, content)
    return jsonify({"ok": True})


@app.route("/api/type/<int:type_id>/rules", methods=["POST"])
def api_add_rule(type_id):
    data = request.get_json()
    content = data.get("content", "")
    if content:
        rule_id = storage.add_rule(type_id, content)
        return jsonify({"ok": True, "rule_id": rule_id})
    return jsonify({"error": "content required"}), 400


@app.route("/api/dataset/<int:dataset_id>/rows", methods=["GET"])
def api_dataset_rows(dataset_id):
    only_unprocessed = request.args.get("unprocessed") == "1"
    rows = storage.get_data_rows(dataset_id, only_unprocessed=only_unprocessed)
    result = []
    for r in rows:
        result.append({
            "row_id": r["row_id"],
            "data": json.loads(r["row_data"]),
            "processed": r["processed"],
            "result": r["result"],
        })
    return jsonify(result)


@app.route("/api/row/<int:row_id>/result", methods=["PUT"])
def api_update_row(row_id):
    data = request.get_json()
    result = data.get("result", "")
    storage.update_row_result(row_id, result)
    return jsonify({"ok": True})


if __name__ == "__main__":
    app.run(debug=True, port=5050)
