"""
Excel 导入导出

导入格式（客服录入表）:
    | 订单号 | 事件类型 | 客户姓名 | 电话 | 地址 | 商品信息 | 新地址 | 备注 |

    事件类型：新增订单 / 恢复订单 / 中断订单 / 修改地址 / 取消订单
    - 新增订单：必填 客户姓名、电话、地址、商品信息
    - 修改地址：必填 新地址
    - 其他类型：可选填 备注

导出格式（交给下一环节的发货表）:
    | 订单号 | 客户姓名 | 电话 | 地址 | 商品信息 | 订单状态 | 批次 |
"""
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import (
    EVENT_TYPE_LABELS,
    EVENT_TYPE_LABELS_REV,
    ORDER_STATUS_LABELS,
    EventType,
    Order,
)
from .storage import Storage

# ── 导入列定义 ──
IMPORT_HEADERS = ["订单号", "事件类型", "客户姓名", "电话", "地址", "商品信息", "新地址", "备注"]

# ── 导出列定义 ──
EXPORT_HEADERS = ["订单号", "客户姓名", "电话", "地址", "商品信息", "订单状态", "批次"]


def generate_import_template(output_path: str | Path):
    """生成空白的客服录入 Excel 模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "客服录入"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col, header in enumerate(IMPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 设置列宽
    widths = [14, 12, 12, 15, 30, 30, 30, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    # 添加示例数据
    examples = [
        ["ORD-001", "新增订单", "张三", "13800001111", "上海市浦东新区XX路100号", "有机苹果x5, 土鸡蛋x30", "", ""],
        ["ORD-002", "新增订单", "李四", "13900002222", "北京市朝阳区YY街200号", "新鲜草莓x3", "", ""],
        ["ORD-002", "修改地址", "", "", "", "", "北京市海淀区新地址888号", "客户要求改地址"],
        ["ORD-001", "中断订单", "", "", "", "", "", "客户暂时联系不上"],
        ["ORD-001", "恢复订单", "", "", "", "", "", "客户已确认"],
        ["ORD-003", "取消订单", "", "", "", "", "", "客户不想要了"],
    ]
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(str(output_path))
    print(f"模板已生成: {output_path}")


def import_events_from_excel(file_path: str | Path) -> list[dict]:
    """
    从 Excel 读取事件列表，返回标准化的事件字典列表。
    每个字典包含: order_id, event_type(EventType), payload(dict)
    不直接写库，由 service 层统一处理。
    """
    wb = load_workbook(str(file_path), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
    wb.close()

    events = []
    for row_num, row in enumerate(rows, 2):
        if not row or not row[0]:
            continue  # 跳过空行

        order_id = str(row[0]).strip()
        event_type_str = str(row[1]).strip() if row[1] else ""

        if event_type_str not in EVENT_TYPE_LABELS:
            print(f"[WARN] 第{row_num}行: 未知事件类型 '{event_type_str}'，跳过")
            continue

        event_type = EVENT_TYPE_LABELS[event_type_str]

        # 根据事件类型组装 payload
        payload = {}
        if event_type == EventType.ORDER_NEW:
            payload = {
                "customer_name": str(row[2] or "").strip(),
                "address": str(row[4] or "").strip(),
                "phone": str(row[3] or "").strip(),
                "items": _parse_items_text(str(row[5] or "")),
            }
        elif event_type == EventType.AFTER_SALE_ADDRESS:
            payload = {"new_address": str(row[6] or "").strip()}
        else:
            payload = {"reason": str(row[7] or "").strip()}

        events.append({
            "order_id": order_id,
            "event_type": event_type,
            "payload": payload,
        })

    print(f"从 Excel 读取 {len(events)} 条事件")
    return events


def export_orders_to_excel(orders: list[Order], output_path: str | Path):
    """将订单列表导出为 Excel，交给下一个环节"""
    wb = Workbook()
    ws = wb.active
    ws.title = "发货订单"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    widths = [14, 12, 15, 30, 30, 10, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    for row_idx, order in enumerate(orders, 2):
        items_display = _format_items_display(order.items)
        status_label = ORDER_STATUS_LABELS.get(order.status, order.status.value)

        ws.cell(row=row_idx, column=1, value=order.order_id)
        ws.cell(row=row_idx, column=2, value=order.customer_name)
        ws.cell(row=row_idx, column=3, value=order.phone)
        ws.cell(row=row_idx, column=4, value=order.address)
        ws.cell(row=row_idx, column=5, value=items_display)
        ws.cell(row=row_idx, column=6, value=status_label)
        ws.cell(row=row_idx, column=7, value=order.batch_id)

    wb.save(str(output_path))
    print(f"导出完成: {output_path}，共 {len(orders)} 条订单")


def _parse_items_text(text: str) -> list[dict]:
    """解析商品文本，如 '有机苹果x5, 土鸡蛋x30' -> [{"name":"有机苹果","qty":5}, ...]"""
    if not text.strip():
        return []
    items = []
    for part in text.split(","):
        part = part.strip()
        if not part:
            continue
        # 尝试解析 "名称x数量" 或 "名称*数量" 格式
        for sep in ["x", "X", "*", "×"]:
            if sep in part:
                name, qty_str = part.rsplit(sep, 1)
                try:
                    items.append({"name": name.strip(), "qty": int(qty_str.strip())})
                except ValueError:
                    items.append({"name": part, "qty": 1})
                break
        else:
            items.append({"name": part, "qty": 1})
    return items


def _format_items_display(items_json: str) -> str:
    """将 JSON 商品信息格式化为可读文本"""
    try:
        items = json.loads(items_json)
        if isinstance(items, list):
            return ", ".join(f'{it["name"]}x{it["qty"]}' for it in items if "name" in it)
    except (json.JSONDecodeError, KeyError):
        pass
    return items_json
